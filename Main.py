import openpyxl 
import csv
import array
import discord
workbook = openpyxl.load_workbook('BotC-Stats.xlsx')
sheet = workbook.active
def main() -> None: 
    client = setupDiscord()
    @client.event
    async def on_ready():
        print(f'Running as {client.user}')

    @client.event
    async def on_message(message):
        checkMessage(client,message)


    client.run('MTMwMzcxOTAwOTc5NTA0NzUyNg.Go9fCh.HTWyggF4jheVsZLmvOFhkdPW7TAuzI1p-jfmCI')

def setupDiscord() -> discord.client:
    intents = discord.Intents.default()
    intents.message_content = True
    client = discord.Client(intents=intents)
    return client
    
def checkMessage(client: discord.client,message: discord.message) -> None:
    if message.author == client.user:
            return
    if message.content.startswith('!PersonalAverageStats'):
        PersonalAverage()

async def PersonalAverage(client: discord.client,message: discord.message)-> None:
    Poster = message.author
    SearchColumn = GetPlayerFromDiscord(Poster.name)
    TotalGood = sheet.cell(row=102, column=SearchColumn).value
    TotalEvil = sheet.cell(row=156, column=SearchColumn).value
    Total = sheet.cell(row=157, column=SearchColumn).value
    await message.channel.send('Your average winrate is ' + Total + ' consisting of ' + TotalGood + ' whilst good and ' + TotalEvil + ' whilst evil!')

def GetPlayerFromDiscord(name: str) -> str:
    match name:
        case "slane3470":
            return 9
        case ".celari":
            return 14
        case "dadude":
            return 19
        case "draconic_lord":
            return 24
        case "thereligionofpeanut":
            return 29
        case "orourkustortoise":
            return 34
        case "toomai1970":
            return 39
        case "lazyvult":
            return 44 
        case "antinium1312.":
            return 49
        case "brotatornator666":
            return 54
        case "ianoid":
            return 59
        case "roobinski":
            return 64
        case "bossors":
            return 69
        case "tsarplatinum":
            return 74
        case "ordainedtick266":
            return 79
        case "hamsternaut":
            return 84
        case ".defize":
            return 89
        case "._._neon_._.":
            return 94
        case "vandyss":
            return 99
        case "brob5046":
            return 104
        case "benign_skies":
            return 109
        case "cgotnr":
            return 114
        case "seventyseven_77":
            return 119
        case "trees17":
            return 124
        case "jetotavio":
            return 129
        case "livburrowss":
            return 134
        case "sincerenumber82":
            return 139
        case "withasideofsalt":
            return 144
        case "thorijus":
            return 149
        case _: #Error if not found player
            return 0
        
def FindPlayer(player: str) -> int:
    match player:
        case "Total":
            return 4
        case "Oliver":
            return 9
        case "Sophie":
            return 14
        case "Ethan":
            return 19
        case "Richard":
            return 24
        case "Dan":
            return 29
        case "Callum":
            return 34
        case "Ryan":
            return 39
        case "Stuart":
            return 44 
        case "Ant":
            return 49
        case "Ronnie":
            return 54
        case "Ian":
            return 59
        case "Reuben":
            return 64
        case "Findlay":
            return 69
        case "Daniel":
            return 74
        case "Emily":
            return 79
        case "Laurence":
            return 84
        case "BenN":
            return 89
        case "BenS":
            return 94
        case "Andy":
            return 99
        case "Ben3":
            return 104
        case "Carrick":
            return 109
        case "Connor":
            return 114
        case "Drystan":
            return 119
        case "Heather":
            return 124
        case "Etienne":
            return 129
        case "Liv":
            return 134
        case "Rory":
            return 139
        case "Scott":
            return 144
        case "TJ":
            return 149
        case _: #Error if not found player
            return 0
        
def FindRole(Role: str) -> int:
    match Role:
        #Townsfolk
        case "Acrobat":
            return 5
        case "Alchemist":
            return 6
        case "Alsaahir":
            return 7
        case "Amnesiac":
            return 8
        case "Artist":
            return 9
        case "Atheist":
            return 10
        case "Ballonist":
            return 11
        case "Banshee":
            return 12
        case "Bounty Hunter":
            return 13
        case "Cannibal":
            return 14
        case "Chambermaid":
            return 15
        case "Chef":
            return 16
        case "Choirboy":
            return 17
        case "Clockmaker":
            return 18
        case "Courtier":
            return 19
        case "Cult Leader":
            return 20
        case "Dreamer":
            return 21
        case "Empath":
            return 22
        case "Engineer":
            return 23
        case "Exorcist":
            return 24
        case "Farmer":
            return 25
        case "Fisherman":
            return 26
        case "Flowergirl":
            return 27
        case "Fool":
            return 28
        case "Fortune Teller":
            return 29
        case "Gambler":
            return 30
        case "General":
            return 31
        case "Gossip":
            return 32
        case "Grandmother":
            return 33
        case "High Priestess":
            return 34
        case "Huntsman":
            return 35
        case "Innkeeper":
            return 36
        case "Investigator":
            return 37
        case "Juggler":
            return 38
        case "King":
            return 39
        case "Knight":
            return 40
        case "Librarian":
            return 41
        case "Lycanthrope":
            return 42
        case "Magician":
            return 43
        case "Mathmetician":
            return 44
        case "Mayor":
            return 45
        case "Minstrel":
            return 46
        case "Monk":
            return 47
        case "Nightwatchman":
            return 48
        case "Noble":
            return 49
        case "Oracle":
            return 50
        case "Pacifist":
            return 51
        case "Philosopher":
            return 52
        case "Pixie":
            return 53
        case "Poppy Grower":
            return 54
        case "Preacher":
            return 55
        case "Professor":
            return 56
        case "Ravenkeeper":
            return 57
        case "Sage":
            return 58
        case "Sailor":
            return 59
        case "Savant":
            return 60
        case "Seamstress":
            return 61
        case "Shugenja":
            return 62
        case "Slayer":
            return 63
        case "Snake Charmer":
            return 64
        case "Soldier":
            return 65
        case "Steward":
            return 66
        case "Tea Lady":
            return 67
        case "Town Crier":
            return 68
        case "Undertaker":
            return 69
        case "Village Idiot":
            return 70
        case "Virgin":
            return 71
        case "Washerwoman":
            return 72
        case "Townsfolk":
            return 73
        #Outsiders
        case "Barber":
            return 77
        case "Butler":
            return 78
        case "Damsel":
            return 79
        case "Drunk":
            return 80
        case "Golem":
            return 81
        case "Goon":
            return 82
        case "Hatter":
            return 83
        case "Heretic":
            return 84
        case "Klutz":
            return 85
        case "Lunatic":
            return 86
        case "Moonchild":
            return 87
        case "Mutant":
            return 88
        case "Ogre":
            return 89
        case "Plague Doctor":
            return 90
        case "Politician":
            return 91
        case "Puzzlemaster":
            return 92
        case "Recluse":
            return 93
        case "Saint":
            return 94
        case "Snitch":
            return 95
        case "Sweetheart":
            return 96
        case "Tinker":
            return 97
        case "Zealot":
            return 98
        case "Outsider":
            return 99
        #Minions
        case "Assassin":
            return 106
        case "Baron":
            return 107
        case "Boffin":
            return 108
        case "Boomdandy":
            return 109
        case "Cerenovus":
            return 110
        case "Devil's Advocate":
            return 111
        case "Evil Twin":
            return 112
        case "Fearmonger":
            return 113
        case "Goblin":
            return 114
        case "Godfather":
            return 115
        case "Harpy":
            return 116
        case "Marionette":
            return 117
        case "Mastermind":
            return 118
        case "Mezepheles":
            return 119
        case "Organ Grinder":
            return 120
        case "Pit-Hag":
            return 121
        case "Poisoner":
            return 122
        case "Psychopath":
            return 123
        case "Scarlet Woman":
            return 124
        case "Spy":
            return 125
        case "Summoner":
            return 126
        case "Vizier":
            return 127
        case "Widow":
            return 128
        case "Witch":
            return 129
        case "Minion":
            return 130
        #Demons
        case "Al-Hadikhia":
            return 134
        case "Fang Gu":
            return 134
        case "Imp":
            return 134
        case "Kazali":
            return 134
        case "Legion":
            return 134
        case "Leviathan":
            return 134
        case "Lil'Monsta":
            return 134
        case "Lleech":
            return 134
        case "Lord of Typhon":
            return 134
        case "No Dashii":
            return 134
        case "Ojo":
            return 134
        case "Po":
            return 134
        case "Pukka":
            return 134
        case "Riot":
            return 134
        case "Shabaloth":
            return 134
        case "Vigormortis":
            return 134
        case "Vortox":
            return 134
        case "Yaggababble":
            return 134
        case "Zombuul":
            return 134
        case "Demon":
            return 135
        #Totals
        case "Total Good":
            return 102
        case "Total Evil":
            return 156
        case "Total":
            return 157
        #Error case if not found
        case _:
            return 0
        
        
def separateFile() -> array:
    with open('results.csv', newline='') as csvfile:
        reader = csv.reader(csvfile, delimiter=' ', quotechar='|')
        Data = array('w')
        for row in reader:
            Data.append(row)
        return Data

def replacePlayerArray(Player: array) ->array:
    FixedPlayer = array('L')
    for entry in Player:
        FixedPlayer.append(FindPlayer(entry))
    return FixedPlayer

def replaceRoleArray(Role: array) ->array:
    FixedRole = array('L')
    for entry in Role:
        FixedRole.append(FindRole(entry))
    return FixedRole

def updateGoodStat(searchcolumn: int, searchrow: int, good_win: bool) -> None:
    if(not(good_win)): #since good has not won increment lost instead of win
        searchcolumn += 1
    searchedcell = sheet.cell(row = searchrow, column= searchcolumn)
    searchedcell.value = searchedcell.value + 1

def updateEvilStat(searchcolumn: int, searchrow: int, good_win: bool) -> None:
    if(good_win): #since good has won evil has not and thus must increment lost instead
        searchcolumn += 1
    searchedcell = sheet.cell(row = searchrow, column= searchcolumn)
    searchedcell.value = searchedcell.value + 1

def updateStats(Data: array) -> None:
    i = 0
    Player = array('w')
    Role = array('w')
    for entry in Data:
        if(i == 0):
            good_win = entry
        elif(i % 2 == 0):
            Player.append()
        else:
            Role.append()
        i += 1
    i = 0
    searchcolumn = replacePlayerArray
    searchrow = replaceRoleArray
    for entry in searchrow:
        if(searchrow[i] >= 106):
            updateEvilStat(searchcolumn[i],searchrow[i],good_win)
        else:
            updateGoodStat(searchcolumn[i],searchrow[i],good_win)
        i += 1

def getPlayerStats(Player: str, Role: str):
    searchcolumn = FindPlayer(Player)
    searchrow = FindRole(Role)
    searchedcell = sheet.cell(row = searchrow, column= searchcolumn)

def saveAndClose() -> None:
    workbook.save('BotC-Stats.xlsx')
    
