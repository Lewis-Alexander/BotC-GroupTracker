import array
import csv
import openpyxl
from openpyxl.utils import get_column_letter
from Spreadsheetclass import spreadsheetValues
from pathlib import Path
import platform
import subprocess

if platform.system() == 'Windows':
    import win32com.client as win32

# Open the workbook in data_only mode
workbook = openpyxl.load_workbook('BotC-Stats.xlsx', data_only=True)
sheet = workbook['Sheet1']

# Workbook for editing (normal mode)
workbook_edit = None
sheet_edit = None

def open_workbook_edit():
    global workbook_edit, sheet_edit
    workbook_edit = openpyxl.load_workbook('BotC-Stats.xlsx')
    sheet_edit = workbook_edit['Sheet1']

def close_workbook_edit():
    global workbook_edit, sheet_edit
    if workbook_edit:
        workbook_edit.close()
        workbook_edit = None
        sheet_edit = None

def recalculate_and_cache_workbook():
    try:
        if platform.system() == 'Windows':
            excel = win32.gencache.EnsureDispatch('Excel.Application')
            excel.Visible = False
            excel.ScreenUpdating = False
            
            # Open the workbook
            workbook = excel.Workbooks.Open(str(Path('BotC-Stats.xlsx').absolute()))
            
            # Force full recalculation
            excel.CalculateFull()
            workbook.Save()
            workbook.Close()
            
            excel.ScreenUpdating = True
            excel.Quit()
        else:
            # Use LibreOffice on Linux/Mac
            import uno
            from com.sun.star.beans import PropertyValue
            
            # Start soffice in listener mode
            subprocess.Popen(['soffice', '--headless', '--accept=socket,host=localhost,port=2002;urp;StarOffice.ServiceManager'], 
                           stdout=subprocess.DEVNULL, stderr=subprocess.DEVNULL)
            
            localContext = uno.getComponentContext()
            resolver = localContext.ServiceManager.createInstanceWithContext(
                "com.sun.star.bridge.UnoUrlResolver", localContext)
            
            try:
                ctx = resolver.resolve(
                    "uno:socket,host=localhost,port=2002;urp;StarOffice.ComponentContext")
                smgr = ctx.ServiceManager
                desktop = smgr.createInstanceWithContext("com.sun.star.frame.Desktop", ctx)
                
                # Open the document
                file_url = uno.systemPathToFileUrl(str(Path('BotC-Stats.xlsx').absolute()))
                doc = desktop.loadComponentFromURL(file_url, "_blank", 0, ())
                
                # Force recalculation
                doc.calculateAll()
                
                # Save the document
                doc.store()
                doc.close(True)
                
            except Exception as e:
                print(f"Warning: Could not recalculate with LibreOffice: {e}")
                
    except Exception as e:
        print(f"Warning: Could not recalculate formulas: {e}")
        pass

def refresh_data_workbook():
    global workbook, sheet
    workbook.close()
    workbook = openpyxl.load_workbook('BotC-Stats.xlsx', data_only=True)
    sheet = workbook['Sheet1']

def separate_file() -> array:
    with open('Results.csv', newline='') as csvfile:
        reader = csv.reader(csvfile, delimiter=',')
        Data = []
        for row in reader:
            Data.append(row)
        return Data

def replace_player_array(Player: array) ->array:
    FixedPlayer = []
    for entry in Player:
        player = entry[0]
        player = player.lower()
        FixedPlayer.append(find_player(player))
    return FixedPlayer

def replace_role_array(Role: array) ->array:
    FixedRole = []
    for entry in Role:
        role = entry[0]
        role = role.lower()
        FixedRole.append(find_role(role))
    return FixedRole

def update_good_stat(column: int, row: int, good_win: int) -> None:
    if good_win[0] == '0':  # since good has not won increment lost instead of win
        column += 1
    cell = f"{get_column_letter(column)}{row}"
    value = sheet_edit[cell].value or 0
    value += 1
    sheet_edit[cell].value = value
    workbook_edit.save('BotC-Stats.xlsx')
    

def update_evil_stat(column: int, row: int, good_win: int) -> None:
    if good_win[0] == '1':  # since good has won evil has not and thus must increment lost instead
        column += 1
    cell = f"{get_column_letter(column)}{row}"
    value = sheet_edit[cell].value or 0
    value += 1
    sheet_edit[cell].value = value
    workbook_edit.save('BotC-Stats.xlsx')
    
def update_player_matchup(column: int, row: int, col_player: int, row_player: int,won: int) -> None:
    if((col_player == 1 and (row_player == 2 or row_player == 3)) or ((col_player == 2 or col_player == 3) and row_player == 1)): #on different teams
        return
    if((won[0] == '0' and col_player == 1) or (won[0] == '1' and (col_player == 2 or col_player == 3))): #since player has not won increment lost instead of win
        column += 1
    gap = 0
    evil = 0 #to change total evil matchups aswell
    if(col_player == 1 and row_player == 1): #both town
        gap = 5
    elif(col_player == 2 and row_player == 2): # both minion
        gap = 1
        evil = 4 - gap
    elif((col_player == 2 and row_player == 3) or (col_player == 3 and row_player == 3)): # col player is minion and row player is demon or both demon
        gap = 2
        evil = 4 - gap
    elif(col_player == 3 and row_player == 2): # col player is demon and row player is minion
        gap = 3
        evil = 4 - gap
    row = row + gap
    cell = f"{get_column_letter(column)}{row}"
    value = sheet_edit[cell].value or 0
    value += 1
    sheet_edit[cell].value = value
    workbook_edit.save('BotC-Stats.xlsx')
    if(evil == 0): # was on good team so no evil matchup to update
        return
    row = row + evil
    cell = f"{get_column_letter(column)}{row}"
    value = sheet_edit[cell].value or 0
    value += 1
    sheet_edit[cell].value = value
    workbook_edit.save('BotC-Stats.xlsx')
    
def update_stats(Data: array) -> None:
    i = 0
    Player = []
    Role = []
    for entry in Data:
        if(i == 0):
            good_win = entry
        elif(i % 2 == 0):
            Role.append(entry)
        else:
            Player.append(entry)
        i += 1
    i = 0
    column = replace_player_array(Player)
    row = replace_role_array(Role)
    matchup_row = replace_player_array_matchup(Player);
    matchup_role = replace_role_array_matchup(row);
    for entry in row:
        j = 0
        if(row[i] >= 106):
            update_evil_stat(column[i],row[i],good_win)
        else:
            update_good_stat(column[i],row[i],good_win)
        for entry in matchup_row:
            if(i != j):
                update_player_matchup(column[i], matchup_row[j], matchup_role[i], matchup_role[j], good_win)
            j += 1
        
        i += 1
        
def update_matchups(Data: array) -> None:
    i = 0
    Player = []
    Role = []
    for entry in Data:
        if(i == 0):
            good_win = entry
        elif(i % 2 == 0):
            Role.append(entry)
        else:
            Player.append(entry)
        i += 1
    i = 0
    column = replace_player_array(Player)
    row = replace_role_array(Role)
    matchup_row = replace_player_array_matchup(Player);
    matchup_role = replace_role_array_matchup(row);
    for entry in row:
        j = 0
        for entry in matchup_row:
            if(matchup_row[i] != "ERROR" and matchup_role[i] != "ERROR" and i != j):
                update_player_matchup(column[i], matchup_row[j], matchup_role[i],matchup_role[j], good_win)
            j += 1
        i += 1
            

def replace_player_array_matchup(Player: array) ->array:
    FixedPlayer = []
    for entry in Player:
        player = entry[0]
        player = player.lower()
        FixedPlayer.append(find_player_matchup(player))
    return FixedPlayer

def replace_role_array_matchup(Role: array) ->array:
    FixedRole = []
    for entry in Role:
        role = entry
        FixedRole.append(find_role_matchup(role))
    return FixedRole
        
def setup_class():
    column = 1
    row = 1
    spreadsheetValues.player_list.clear()
    spreadsheetValues.player_col_list.clear()
    spreadsheetValues.username_list.clear()
    spreadsheetValues.role_list.clear()
    spreadsheetValues.role_list_idx.clear()
    while True:
        cell = f"{get_column_letter(column)}{row}"
        value = sheet[cell].value
        if value == 'end':
            break
        elif value and value not in ['Players', 'Total', 'Template']:
            spreadsheetValues.player_list.append(value.lower())
            spreadsheetValues.player_col_list.append(column)
            row += 1
            cell = f"{get_column_letter(column)}{row}"
            value = sheet[cell].value
            if value:
                spreadsheetValues.username_list.append(value.lower())
            row -= 1
        column += 1
    spreadsheetValues.playercount = len(spreadsheetValues.player_list)
    
    column = 1
    while(True):
        cell = f"{get_column_letter(column)}{row}"
        value = sheet[cell].value
        match value:
            case 'roleend':
                row += 1
                break
            case None:
                pass
            case 'Townsfolk Total':
                spreadsheetValues.townsfolk = row
            case 'Outsider Total':
                spreadsheetValues.outsider = row
            case 'Minion Total':
                spreadsheetValues.minion = row
            case 'Demon Total':
                spreadsheetValues.demon = row
            case 'Good Averages':
                spreadsheetValues.average_good = row
            case 'Evil Averages':
                spreadsheetValues.average_evil = row
            case 'Total Averages':
                spreadsheetValues.average_total = row
            case 'Total Played':
                spreadsheetValues.total_played = row
            case _ if value not in ['Players','Usernames','Townsfolk Roles', 'Outsider Roles', 'Minion Roles', 'Demon Roles']:
                spreadsheetValues.role_list.append(value.lower())
                spreadsheetValues.role_list_idx.append(row)
        row += 1
    spreadsheetValues.rolecount = len(spreadsheetValues.role_list)
    spreadsheetValues.matchup_row_start = row
    while(True):
        cell = f"{get_column_letter(column)}{row}"
        value = sheet[cell].value
        match value:
            case None:
                pass
            case _ if value.lower() == spreadsheetValues.player_list[1]: #since the gap between players is constant just need to find first
                break
        row += 1
    spreadsheetValues.matchup_gap = row - spreadsheetValues.matchup_row_start   


def find_role(rolein: str) -> str:
    match rolein.lower():
        case "townsfolk":
            return spreadsheetValues.townsfolk
        case "outsider":
            return spreadsheetValues.outsider
        case "minion":
            return spreadsheetValues.minion
        case "demon":
            return spreadsheetValues.demon
        case "total good":
            return spreadsheetValues.average_good
        case "total evil":
            return spreadsheetValues.average_evil
        case "total":
            return spreadsheetValues.average_total
        case _:
            for role in spreadsheetValues.role_list:
                if(role.lower() == rolein.lower()):
                    index = spreadsheetValues.role_list.index(role)
                    return spreadsheetValues.role_list_idx[index]
    return "ERROR"
        
def find_player(playerin: str) -> str:
    for player in spreadsheetValues.player_list:
        if(player.lower() == playerin.lower()):
            index = spreadsheetValues.player_list.index(player)
            return spreadsheetValues.player_col_list[index]
    return "ERROR"

def find_player_username(name: str) -> str:
    for player in spreadsheetValues.username_list:
        if(player.lower() == name):
            index = spreadsheetValues.username_list.index(player)
            return spreadsheetValues.player_col_list[index]
    return "ERROR"

def find_player_matchup(name: str) -> str:
    for player in spreadsheetValues.player_list:
        if(player.lower() == name):
            index = spreadsheetValues.player_list.index(player)
            return spreadsheetValues.matchup_row_start + (index * spreadsheetValues.matchup_gap)
    return "ERROR"


def find_role_matchup(Row: int) -> int:
    if(Row <= spreadsheetValues.outsider): #townsfolk or outsider
        return 1
    elif(Row <= spreadsheetValues.minion): #minion
        return 2
    elif(Row <= spreadsheetValues.demon): #demon
        return 3
    else: #error
        return "ERROR"
    
def get_role_image(role: str):
    sanitized_role = role.replace(" ", "").replace("-", "").replace("'", "").lower()
    image_path = Path("role-images") / f"{sanitized_role}.png"
    return image_path if image_path.exists() else None

def get_role_rules(role: str):
    sanitized_role = role.replace(" ", "").replace("-", "").replace("'", "").lower()
    rules_path = Path("role-rules") / f"{sanitized_role}.txt"
    return rules_path if rules_path.exists() else None