import discord
import Helper
import Pairwise
import random
import traceback
from Token import token, bot_channel_id, server_id, role_0_id, role_10_id, role_20_id, role_40_id, role_60_id, role_80_id, role_100_id, error_user_id
from discord import app_commands
from discord.ui import View, Button
from discord.ext import commands
from pathlib import Path
from openpyxl.utils import get_column_letter
from Spreadsheetclass import spreadsheetValues
import csv
import openpyxl

bot = commands.Bot(command_prefix='!', intents = discord.Intents.all())

async def send_error_to_discord(error: Exception, context: str = ""):
    """Send an error message to the specified Discord user."""
    if error_user_id is None:
        print(f"Error notification not configured. error_user_id is not set in Token.py")
        print(f"Error in {context}: {str(error)}")
        print(traceback.format_exc())
        return
    
    try:
        user = await bot.fetch_user(error_user_id)
        error_message = f"**Error occurred{' in ' + context if context else ''}:**\n\n"
        error_message += f"```\n{str(error)}\n```\n"
        error_message += f"**Traceback:**\n```\n{traceback.format_exc()}\n```"
        
        # Discord has a 2000 character limit, so split if necessary
        if len(error_message) > 2000:
            # Send the main error message
            await user.send(error_message[:2000])
            # Send the rest
            await user.send(error_message[2000:])
        else:
            await user.send(error_message)
    except Exception as e:
        print(f"Failed to send error notification to Discord user: {str(e)}")
        print(f"Original error: {str(error)}")

@bot.event
async def on_app_command_error(interaction: discord.Interaction, error: app_commands.AppCommandError):
    """Global error handler for app commands."""
    await send_error_to_discord(error, f"Command: {interaction.command.name if interaction.command else 'Unknown'}")
    
    # Respond to the user if possible
    try:
        if not interaction.response.is_done():
            await interaction.response.send_message(
                f"An error occurred while processing your command. The error has been logged.",
                ephemeral=True
            )
        else:
            await interaction.followup.send(
                f"An error occurred while processing your command. The error has been logged.",
                ephemeral=True
            )
    except:
        pass

@bot.event
async def on_error(event, *args, **kwargs):
    """Global error handler for events."""
    error = Exception(f"Error in event: {event}")
    await send_error_to_discord(error, f"Event: {event}")

@bot.event
async def on_ready():
    print(f'Running as {bot.user}')
    print(bot.user.id)

@bot.command()
@commands.guild_only()
@commands.is_owner()  # Prevent other people from using the command
async def sync(ctx: commands.Context) -> None:
    """Sync app commands to Discord."""
    try:
        synced = await bot.tree.sync()
        print(f"Synced {len(synced)} command(s)")
    except Exception as e:
        print(e)
    

@bot.tree.command(name="personal_average", description="Check your average winrates for sides")
async def personal_average(interaction: discord.Interaction):
    column = Helper.find_player_username(interaction.user.name)
    column += 2
    cell = f"{get_column_letter(column)}{spreadsheetValues.average_good}"
    TotalGood = Helper.sheet[cell].value
    cell = f"{get_column_letter(column)}{spreadsheetValues.average_evil}"
    TotalEvil = Helper.sheet[cell].value
    cell = f"{get_column_letter(column)}{spreadsheetValues.average_total}"
    Total = Helper.sheet[cell].value
    cell = f"{get_column_letter(column)}{spreadsheetValues.total_played}"
    Played = Helper.sheet[cell].value
    await interaction.response.send_message(f"Over {str(Played)} games played your average winrate was {str(Total)} consisting of {str(TotalGood)} whilst good and {str(TotalEvil)} whilst evil!", ephemeral=True)



@bot.tree.command(name="personal_role_stats", description="Check any your stats for a particular role")
@app_commands.describe(role = 'Role you would like to check (Townsfolk for Townsfolk total and Total Good for all good)')
async def personal_role_stats(interaction: discord.Interaction, role: str):
    column = Helper.find_player_username(interaction.user.name)
    column -= 1
    row = Helper.find_role(role.lower())
    if(row == 0):
        await interaction.response.send_message(f'Role not found please check spelling', ephemeral=True)
    else:
        data = []
        for i in range(4):     
            cell = f"{get_column_letter(column)}{row}"
            data.append(Helper.sheet[cell].value)
            column += 1
        await interaction.response.send_message(f'you have played the {role.capitalize()} {data[0]} times, of those you won {data[1]} and lost {data[2]} which makes your winrate {data[3]}.')

@bot.tree.command(name="player_average", description="Check any players average winrates")
@app_commands.describe(player = 'Player you would like to see the averages for')
async def player_average(interaction: discord.Interaction, player: str):
    column = Helper.find_player(player.lower())
    if(column == "ERROR"):
        await interaction.response.send_message(f'Player not found please check spelling', ephemeral=True)
    else:
        column += 2
        data = []
        rows = [spreadsheetValues.average_good,spreadsheetValues.average_evil,spreadsheetValues.average_total]
        for i in range(3):       
            cell = f"{get_column_letter(column)}{rows[i]}"
            data.append(Helper.sheet[cell].value)
        await interaction.response.send_message(f'{player.capitalize()}\'s average winrates are as follows: Good-{data[0]} Evil-{data[1]} Total-{data[2]}')


@bot.tree.command(name="role_total_stats", description="Check the stats for a particular role")
@app_commands.describe(role = 'Role you would like to check, (Townsfolk for Townsfolk total and Total Good for all good)')
async def role_total_stats(interaction: discord.Interaction, role: str):
    row = Helper.find_role(role.lower())
    if(row == "ERROR"):
        await interaction.response.send_message(f'Role not found please check spelling', ephemeral=True) 
    else: 
        column = 'C'
        data = []
        for i in range(4):       
                cell = f"{get_column_letter(column)}{row}"
                data.append(Helper.sheet[cell].value)
                column += 1
        await interaction.response.send_message(f'The {role.capitalize()} has been played {data[0]} times of those they won {data[1]} and lost {data[2]} which makes their winrate {data[3]}.')

@bot.tree.command(name='player_role_stats', description="Check any players stats for a particular role")
@app_commands.describe(role = 'Role you would like to check, (Townsfolk for Townsfolk total and Total Good for all good)')
@app_commands.describe(player = 'Player you would like to check (same name as on spreadsheet)')
async def player_role_stats(interaction: discord.Interaction, role: str, player: str):
    row = Helper.find_role(role.lower())
    column = Helper.find_player(player.lower())
    if(row == 0):
        await interaction.response.send_message(f'Role not found please check spelling', ephemeral=True)
    elif(column == "ERROR"):
        await interaction.response.send_message(f'Player not found please check spelling', ephemeral=True)
    else:
        data = []
        for i in range(4):       
                cell = f"{get_column_letter(column)}{row}"
                data.append(Helper.sheet[cell].value)
                column += 1
        await interaction.response.send_message(f'{player.capitalize()} has played {role.capitalize()} {data[0]} times of those they won {data[1]} and lost {data[2]} which makes their winrate {data[3]}.', ephemeral=True)   
    
        

@bot.tree.command(name="upload_spreadsheet", description="Uploads the current entire spreadsheet to be perused at your pleasure")
async def upload_spreadsheet(interaction: discord.Interaction):
    await interaction.response.send_message(file=discord.File(r'BotC-Stats.xlsx'))

@commands.is_owner()  # Prevent other people from using the command
@bot.tree.command(name="update_spreadsheet", description="if program has a csv file it uses it to update the spreadsheet")
async def update_spreadsheet(interaction: discord.Interaction):
    await interaction.response.defer(ephemeral=True)
    my_file = Path("Results.csv")
    if my_file.is_file():
        Helper.open_workbook_edit()

        data = Helper.separate_file()

        invalid_entries = []
        for index, entry in enumerate(data):
            if index == 0:
                continue
            else:
                if index % 2 == 0:
                    if entry[0].lower() not in spreadsheetValues.role_list:
                        invalid_entries.append(f"Role: {entry}")
                else:
                    if entry[0].lower() not in spreadsheetValues.player_list:
                        invalid_entries.append(f"Player: {entry}")
                        
        if invalid_entries:
            error_message = "Invalid entries found: " + ", ".join(invalid_entries)
            await interaction.followup.send(error_message, ephemeral=True)
            # Close the edit workbook without saving
            Helper.close_workbook_edit()
            return

        Helper.update_stats(data)
        Helper.close_workbook_edit()
        Helper.recalculate_and_cache_workbook()
        Helper.refresh_data_workbook()

        await interaction.followup.send(f'spreadsheet updated', ephemeral=True)

@commands.is_owner()  # Prevent other people from using the command
@bot.tree.command(name="update_matchups", description="if program has a csv file it uses it to update the spreadsheet")
async def update_matchups(interaction: discord.Interaction):
    await interaction.response.defer(ephemeral=True)
    my_file = Path("Results.csv")
    if my_file.is_file():
        Helper.open_workbook_edit()

        data = Helper.separate_file()
        Helper.update_matchups(data)
        Helper.close_workbook_edit()
        Helper.recalculate_and_cache_workbook()
        Helper.refresh_data_workbook()

        await interaction.followup.send(f'matchups updated', ephemeral=True)      

@bot.tree.command(name="update_role", description="auto update personal role from database")
async def update_role(interaction : discord.Interaction):
    server = bot.get_guild(server_id)
    role0 = server.get_role(role_0_id)
    role10 = server.get_role(role_10_id)
    role20 = server.get_role(role_20_id)
    role40 = server.get_role(role_40_id)
    role60 = server.get_role(role_60_id)
    role80 = server.get_role(role_80_id)
    role100 = server.get_role(role_100_id)
    member = interaction.user
    column = Helper.find_player_username(member.name)
    column += 2
    cell = f"{get_column_letter(column)}{spreadsheetValues.total_played}"
    cellval = Helper.sheet[cell].value
    if(cellval < 10):
        await member.add_roles(role0)
        await interaction.response.send_message(f'You have been assigned the 0+ games role as you have not yet attended enough to increase you currently have {cellval} games played.', ephemeral=True)
    elif(cellval < 20):
        await member.add_roles(role10)
        await member.remove_roles(role0)
        await interaction.response.send_message(f'You have been assigned the 10+ games role as you have attended enough to increase you currently have {cellval} games played.', ephemeral=True)
    elif(cellval < 40):
        await member.add_roles(role20)
        await member.remove_roles(role10)
        await interaction.response.send_message(f'You have been assigned the 20+ games role as you have attended enough to increase you currently have {cellval} games played.', ephemeral=True)
    elif(cellval < 60):
        await member.add_roles(role40)
        await member.remove_roles(role20)
        await interaction.response.send_message(f'You have been assigned the 40+ games role as you have attended enough to increase you currently have {cellval} games played. ', ephemeral=True)
    elif(cellval < 80):
        await member.add_roles(role60)
        await member.remove_roles(role40)
        await interaction.response.send_message(f'You have been assigned the 60+ games role as you have attended enough to increase you currently have {cellval} games played.', ephemeral=True)
    elif(cellval < 100):
        await member.add_roles(role80)
        await member.remove_roles(role60)
        await interaction.response.send_message(f'You have been assigned the 80+ games role as you have attended enough to increase you currently have {cellval} games played.', ephemeral=True)
    else:
        await member.add_roles(role100)
        await member.remove_roles(role80)
        await interaction.response.send_message(f'You have been assigned the 100+ games role as you have attended enough to increase you currently have {cellval} games played.', ephemeral=True)
        
@bot.tree.command(name="highest_role_winrate", description="highest winrate for each role")
@app_commands.describe(role = 'Role you would like to check (Townsfolk for Townsfolk total and Total Good for all good)')
async def highest_role_winrate(interaction : discord.Interaction, role : str):
    column = spreadsheetValues.starting_player_percentage_column
    row = Helper.find_role(role.lower())
    if(row == 0):
        await interaction.response.send_message(f'Role not found please check spelling', ephemeral=True)
    data = []
    for player in range (spreadsheetValues.playercount):
        cell = f"{get_column_letter(column)}{row}"
        tempdata = Helper.sheet[cell].value
        tempdata = tempdata[:-1]
        data.append(float(tempdata))
        for i in range(5):
            column += 1
    unsorteddata = data.copy()
    data.sort(reverse=True)
    i = 0
    highestindex = []
    for entry in data:
        if(data[0] == unsorteddata[i]):
            highestindex.append(i)
        i += 1
    name = []
    for value in highestindex:
        column = spreadsheetValues.starting_player_win_column
        for index in range(value):
            for i in range(5):
                column += 1
        cell = f"{get_column_letter(column)}{1}"
        name.append(Helper.sheet[cell].value)
    if(len(name) > 1):
        players = f'Multiple players have a tied winrate of {data[0]}% on the {role.capitalize()} those being '
        for entry in name:
            if(entry != name[-1]):
                players += f'{entry}, '
            else:
                players += f'and {entry}.'
        await interaction.response.send_message(f'{players}')
    else:
        await interaction.response.send_message(f'{name[0].capitalize()} has the highest winrate as the/a {role.capitalize()} which is {data[0]}%')
    
@commands.is_owner()  # Prevent other people from using the command
@bot.tree.command(name="update_user_role", description="Update the role for a given user")
@app_commands.describe(user="The user whose role you want to update")
async def update_user_role(interaction: discord.Interaction, user: discord.Member):
    try:
        # Retrieve the column for the user
        column = Helper.find_player_username(user.name)
        column += 2
        cell = f"{get_column_letter(column)}{spreadsheetValues.total_played}"
        cellval = Helper.sheet[cell].value

        # Define roles based on the algorithm in update_role
        server = interaction.guild
        role0 = server.get_role(role_0_id)
        role10 = server.get_role(role_10_id)
        role20 = server.get_role(role_20_id)
        role40 = server.get_role(role_40_id)
        role60 = server.get_role(role_60_id)
        role80 = server.get_role(role_80_id)
        role100 = server.get_role(role_100_id)

        # Assign roles based on the total games played
        if cellval < 10:
            await user.add_roles(role0)
            await interaction.response.send_message(f"{user.mention} has been assigned the 0+ games role. they have {cellval} games played.", ephemeral=True)
        elif cellval < 20:
            await user.add_roles(role10)
            await user.remove_roles(role0)
            await interaction.response.send_message(f"{user.mention} has been assigned the 10+ games role. they have {cellval} games played.", ephemeral=True)
        elif cellval < 40:
            await user.add_roles(role20)
            await user.remove_roles(role10)
            await interaction.response.send_message(f"{user.mention} has been assigned the 20+ games role. they have {cellval} games played.", ephemeral=True)
        elif cellval < 60:
            await user.add_roles(role40)
            await user.remove_roles(role20)
            await interaction.response.send_message(f"{user.mention} has been assigned the 40+ games role. they have {cellval} games played.", ephemeral=True)
        elif cellval < 80:
            await user.add_roles(role60)
            await user.remove_roles(role40)
            await interaction.response.send_message(f"{user.mention} has been assigned the 60+ games role. they have {cellval} games played.", ephemeral=True)
        elif cellval < 100:
            await user.add_roles(role80)
            await user.remove_roles(role60)
            await interaction.response.send_message(f"{user.mention} has been assigned the 80+ games role. they have {cellval} games played.", ephemeral=True)
        else:
            await user.add_roles(role100)
            await user.remove_roles(role80)
            await interaction.response.send_message(f"{user.mention} has been assigned the 100+ games role. they have {cellval} games played.", ephemeral=True)
    except Exception as e:
        # Handle any errors that occur
        await interaction.response.send_message(f"Failed to update role: {str(e)}", ephemeral=True)

@bot.tree.command(name="player_to_player_matchup_evil", description="Check two players matchup stats when on the evil team together first is the reference player")
@app_commands.describe(player1 = 'Player you would like to check (same name as on spreadsheet)')
@app_commands.describe(player2 = 'Player you would like to check (same name as on spreadsheet)')
async def player_to_player_matchup_evil(interaction: discord.Interaction, player1: str, player2: str):
    column = Helper.find_player(player1.lower())
    row = Helper.find_player_matchup(player2.lower())
    if(column == "ERROR" or row == "ERROR"):
        await interaction.response.send_message(f'Player not found please check spelling', ephemeral=True)
    else:
        column += 2
        row = row + 1 #start with minion info
        cell = f"{get_column_letter(column)}{row}"
        data = []
        for i in range(4):       
                cell = f"{get_column_letter(column)}{row}"
                data.append(Helper.sheet[cell].value)
                row = row + 1
        await interaction.response.send_message(f'when {player1.capitalize()} and {player2.capitalize()} are both minions their winrate is {data[0]}, if {player2.capitalize()} is the demon or they both are demons their winrate is {data[1]}, if the {player1.capitalize()} is the demon and {player2.capitalize()} is one of the minions their winrate is {data[2]}, this means their average winrate as an evil team is {data[3]}.')
        
@bot.tree.command(name="player_to_player_matchup_good", description="Check two players matchup stats when on the good team together first is the reference player")
@app_commands.describe(player1 = 'Reference Player you would like to check (same name as on spreadsheet)')
@app_commands.describe(player2 = 'Secondary Player you would like to check (same name as on spreadsheet)')
async def player_to_player_matchup_good(interaction: discord.Interaction, player1: str, player2: str):
    column = Helper.find_player(player1.lower())
    row = Helper.find_player_matchup(player2.lower())
    if(column == "ERROR" or row == "ERROR"):
        await interaction.response.send_message(f'Player not found please check spelling', ephemeral=True)
    else:
        column += 2
        row = row + 5 #start with correct info
        cell = f"{get_column_letter(column)}{row}"    
        data = Helper.sheet[cell].value
        row = row + 1
        await interaction.response.send_message(f'when {player1.capitalize()} and {player2.capitalize()} are both on the good team their winrate is {data}.')
        
@bot.tree.command(name="player_to_player_matchup_total", description="Check two players matchup stats when on the same team, first is the reference player")
@app_commands.describe(player1 = 'Reference Player you would like to check (same name as on spreadsheet)')
@app_commands.describe(player2 = 'Secondary Player you would like to check (same name as on spreadsheet)')
async def player_to_player_matchup_total(interaction: discord.Interaction, player1: str, player2: str):
    column = Helper.find_player(player1.lower())
    row = Helper.find_player_matchup(player2.lower())
    if(column == "ERROR" or row == "ERROR"):
        await interaction.response.send_message(f'Player not found please check spelling', ephemeral=True)
    else:
        column += 2
        row = row + 6 #start with correct info
        cell = f"{get_column_letter(column)}{row}"    
        data = Helper.sheet[cell].value
        row = row + 1
        await interaction.response.send_message(f'when {player1.capitalize()} and {player2.capitalize()} are both on the same team their winrate is {data}.')
        

@bot.tree.command(name="player_to_player_winrate_delta", description="Check two players winrate delta when playing together, first is the reference player")
@app_commands.describe(player1 = 'Reference Player you would like to check (same name as on spreadsheet)')
@app_commands.describe(player2 = 'Secondary Player you would like to check (same name as on spreadsheet)')
async def player_to_player_winrate_delta(interaction: discord.Interaction, player1: str, player2: str):
    column = Helper.find_player(player1.lower())
    row = Helper.find_player_matchup(player2.lower())
    if(column == "ERROR" or row == "ERROR"):
        await interaction.response.send_message(f'Player not found please check spelling', ephemeral=True)
    else:
        column += 2
        row = row + 7 #start with correct info
        cell = f"{get_column_letter(column)}{row}"
        data = []
        for i in range(3):       
                cell = f"{get_column_letter(column)}{row}"
                data.append(Helper.sheet[cell].value)
                row = row + 1
        data[2] = data[2]*100
        await interaction.response.send_message(f'when {player1.capitalize()} and {player2.capitalize()} are on the evil team together {player1.capitalize()}\'s default evil winrate changes by {data[0]} (absolute), if they are together on the good team their winrate changes by {data[1]} (absolute), on average if both players are on the same team {player1.capitalize()}\'s average winrate delta is an absolute value of {data[2]}%.')

# reused for both fun and strength comparisons
async def send_new_comparison(interaction: discord.Interaction, is_initial: bool, category: str, message_text: str):
        # Defer the interaction to keep it valid for follow-ups
        if not interaction.response.is_done():
            await interaction.response.defer(ephemeral=True)

        role1, role2 = random.sample(spreadsheetValues.role_list, 2)

        async def button_callback(interaction: discord.Interaction, selected_role: str):
            await interaction.response.edit_message(content="Loading next comparison...", view=None)
            Pairwise.save_pairwise_comparison(role1, role2, selected_role, category=category)
            await send_new_comparison(interaction, is_initial=False, category=category, message_text=message_text)

        async def skip_callback(interaction: discord.Interaction):
            await interaction.response.edit_message(content="Loading next comparison...", view=None)
            await send_new_comparison(interaction, is_initial=False, category=category, message_text=message_text)

        role1_image = Helper.get_role_image(role1)
        role2_image = Helper.get_role_image(role2)

        button1 = Button(label=role1.capitalize(), style=discord.ButtonStyle.primary)
        button2 = Button(label=role2.capitalize(), style=discord.ButtonStyle.primary)
        skip_button = Button(label="Do Not Know", style=discord.ButtonStyle.secondary)

        button1.callback = lambda i: button_callback(i, role1)
        button2.callback = lambda i: button_callback(i, role2)
        skip_button.callback = skip_callback

        view = View()
        view.add_item(button1)
        view.add_item(button2)
        view.add_item(skip_button)

        files = []
        content = message_text
        if role1_image:
            files.append(discord.File(role1_image, filename="role1.png"))
        if role2_image:
            files.append(discord.File(role2_image, filename="role2.png"))

        role1_rules = Helper.get_role_rules(role1)
        role2_rules = Helper.get_role_rules(role2)

        # preventing homebrew since they are non specific
        while "homebrew" in role1.lower() or "homebrew" in role2.lower():
            role1, role2 = random.sample(spreadsheetValues.role_list, 2)

        # Update images and rules after ensuring valid roles
        role1_image = Helper.get_role_image(role1)
        role2_image = Helper.get_role_image(role2)
        role1_rules = Helper.get_role_rules(role1)
        role2_rules = Helper.get_role_rules(role2)

        files = []
        content = message_text
        if role1_image:
            files.append(discord.File(role1_image, filename="role1.png"))
        if role2_image:
            files.append(discord.File(role2_image, filename="role2.png"))

        if role1_rules:
            with open(role1_rules, "r") as file:
                content += f"\n{role1.capitalize()}:\n" + file.read()
        if role2_rules:
            with open(role2_rules, "r") as file:
                content += f"\n{role2.capitalize()}:\n" + file.read()

        if is_initial:
            await interaction.followup.send(
                content=content, view=view, ephemeral=True, files=files
            )
        else:
            await interaction.followup.send(
                content=content, view=view, files=files, ephemeral=True
            )
    
@bot.tree.command(name="role_fun_comparison", description="Randomly select two roles and ask which is more fun to play as/against")
async def role_fun_comparison(interaction: discord.Interaction):
    await send_new_comparison(interaction, is_initial=True, category="fun", message_text="Which role is more fun to play as/against?")
    
@bot.tree.command(name="role_strength_comparison", description="Randomly select two roles and ask which is in your view stronger")
async def role_strength_comparison(interaction: discord.Interaction):
    await send_new_comparison(interaction, is_initial=True, category="strength", message_text="Which role is stronger?")

@bot.tree.command(name="role_ranking", description="Display a ranking of roles based on pairwise comparisons")
@app_commands.choices(category=[
    app_commands.Choice(name="Fun", value="fun"),
    app_commands.Choice(name="Strength", value="strength")
])
async def role_ranking(interaction: discord.Interaction, category: app_commands.Choice[str]):
    ranked_roles = Pairwise.generate_role_ranking(category=category.value)

    if not ranked_roles:
        await interaction.response.send_message("No ranking data available.", ephemeral=True)
        return

    # Export the ranking to a CSV file
    csv_file_path = "role_ranking.csv"
    with open(csv_file_path, "w", newline="", encoding="utf-8") as csvfile:
        writer = csv.writer(csvfile)
        writer.writerow(["Rank", "Role", "Score"])
        for i, (role, score) in enumerate(ranked_roles):
            writer.writerow([i + 1, role.capitalize(), score])

    # Send the CSV file as a response
    await interaction.response.send_message(
        content="The role ranking is too long to display here. Please find the ranking in the attached CSV file.",
        file=discord.File(csv_file_path)
    )
    
@bot.tree.command(name="get_role", description="Get the data for a specific role")
async def get_role(interaction: discord.Interaction, role: str):
    rules_path = Helper.get_role_rules(role)
    image_path = Helper.get_role_image(role)
    files = []
    if image_path:
        files.append(discord.File(image_path, filename="role.png"))

    if rules_path:
        with open(rules_path, "r") as file:
            rules_content = file.read()
        await interaction.response.send_message(
            f"{role.capitalize()}:\n{rules_content}", files=files
        )
    else:
        await interaction.response.send_message(
            f"No role found for: {role}", ephemeral=True, files=files
        )

    
@bot.tree.command(name="upload_all_session_csvs", description="Select a session and upload all CSVs for that session")
async def upload_all_session_csvs(interaction: discord.Interaction):
    sessions_path = Path("Historical Results")
    sessions = [folder.name for folder in sessions_path.iterdir() if folder.is_dir()]

    class SessionDropdown(discord.ui.Select):
        def __init__(self):
            options = [discord.SelectOption(label=session) for session in sessions]
            super().__init__(placeholder="Select a session", min_values=1, max_values=1, options=options)

        async def callback(self, interaction: discord.Interaction):
            selected_session = self.values[0]
            session_path = sessions_path / selected_session
            csv_files = list(session_path.glob("*.csv"))
            if not csv_files:
                await interaction.response.edit_message(content=f"No CSV files found in session {selected_session}.", view=None)
                return

            files = [discord.File(csv_file) for csv_file in csv_files]
            await interaction.response.edit_message(content=f"Uploading CSV files for session {selected_session}, 1 means good team won. 0 means evil team won:", attachments=files, view=None)

    class SessionDropdownView(discord.ui.View):
        def __init__(self):
            super().__init__()
            self.add_item(SessionDropdown())

    await interaction.response.send_message("Please select a session:", view=SessionDropdownView(), ephemeral=True)

@commands.is_owner()  # Prevent other people from using the command    
@bot.tree.command(name="copy_results", description="Copy Results.csv to a selected session directory or create a new session directory")
async def copy_results(interaction: discord.Interaction):
    results_file = Path("Results.csv")
    if not results_file.is_file():
        try:
            if interaction.response.is_done():
                await interaction.followup.send("Results.csv not found.", ephemeral=True)
            else:
                await interaction.response.send_message("Results.csv not found.", ephemeral=True)
        except discord.errors.NotFound:
            await interaction.response.send_message("Results.csv not found.", ephemeral=True)
        return

    sessions_path = Path("Historical Results")
    sessions = [folder.name for folder in sessions_path.iterdir() if folder.is_dir()]

    class SessionDropdown(discord.ui.Select):
        def __init__(self):
            options = [discord.SelectOption(label=session) for session in sessions]
            super().__init__(placeholder="Select a session", min_values=1, max_values=1, options=options)

        async def callback(self, interaction: discord.Interaction):
            selected_session = self.values[0]
            session_path = sessions_path / selected_session

            # Determine and use the session number from the directory name
            session_number = int(selected_session.split('-')[1])
            existing_files = list(session_path.glob(f"S{session_number}Results*.csv"))
            new_file_number = len(existing_files) + 1
            new_file_name = f"S{session_number}Results{new_file_number}.csv"

            destination = session_path / new_file_name
            try:
                destination.write_bytes(results_file.read_bytes())
                await interaction.response.edit_message(content=f"Results.csv has been copied to {selected_session} as {new_file_name}.", view=None)
            except Exception as e:
                await interaction.response.edit_message(content=f"Failed to copy Results.csv: {str(e)}", view=None)

    class CreateSessionButton(discord.ui.Button):
        def __init__(self):
            super().__init__(label="Create New Session", style=discord.ButtonStyle.primary)

        async def callback(self, interaction: discord.Interaction):
            new_session_number = 1
            while (sessions_path / f"Session-{new_session_number}").exists():
                new_session_number += 1
            new_session_path = sessions_path / f"Session-{new_session_number}"
            try:
                new_session_path.mkdir()
                session_number = new_session_number
                new_file_name = f"S{session_number}Results1.csv"
                destination = new_session_path / new_file_name
                destination.write_bytes(results_file.read_bytes())

                await interaction.response.edit_message(content=f"Created new session directory: Session-{new_session_number} and copied Results.csv as {new_file_name}.", view=None)
            except Exception as e:
                await interaction.response.edit_message(content=f"Failed to create new session directory or copy Results.csv: {str(e)}", view=None)

    class SessionDropdownView(discord.ui.View):
        def __init__(self):
            super().__init__()
            self.add_item(SessionDropdown())
            self.add_item(CreateSessionButton())

    try:
        if interaction.response.is_done():
            await interaction.followup.send("Please select a session or create a new one:", view=SessionDropdownView(), ephemeral=True)
        else:
            await interaction.response.send_message("Please select a session or create a new one:", view=SessionDropdownView(), ephemeral=True)
    except discord.errors.NotFound:
        await interaction.response.send_message("Please select a session or create a new one:", view=SessionDropdownView(), ephemeral=True)

@bot.event
async def on_ready():
    Helper.setup_class()
    await bot.get_channel(bot_channel_id).send(f"Bot is running and has loaded all current players and roles from the spreadsheet currently {len(spreadsheetValues.username_list)} players and {spreadsheetValues.rolecount} roles.")


bot.run(token)
